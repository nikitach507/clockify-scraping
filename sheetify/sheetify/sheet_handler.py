import os
import json
import gspread
import time
from googleapiclient.errors import HttpError
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from datetime import datetime
from openpyxl.utils import get_column_letter
import gspread


class GoogleSheetAPI:
    def __init__(self, spreadsheet_id: str, credentials_path: str, token_path: str) -> None:
        self.credentials_path = credentials_path
        self.token_path = token_path
        self.spreadsheet_id = spreadsheet_id
        self.gc = None
        self.credentials = None
        self.service = None
        self.worksheet = None
        self.sheet_id = None

    def _authorize(self):
        scopes = ['https://www.googleapis.com/auth/spreadsheets']

        try:
            if os.path.exists(self.token_path):
                with open(self.token_path, 'r') as token:
                    self.credentials = Credentials.from_authorized_user_info(json.load(token), scopes)
        except (FileNotFoundError, TypeError) as e:
            raise e

        if not self.credentials or not self.credentials.valid:
            if self.credentials and self.credentials.expired and self.credentials.refresh_token:
                self.credentials.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(self.credentials_path, scopes)
                self.credentials = flow.run_local_server(port=0)

            with open(self.token_path, 'w') as token:
                token.write(self.credentials.to_json())

        self.gc = gspread.authorize(self.credentials)
        self.service = build('sheets', 'v4', credentials=self.credentials)

    def _safety_append_rows(self, data: list[list[str]], value_input_option: str = None, row: bool = False) -> None:
        while True:
            try:
                if row:
                    self.worksheet.append_row(data, value_input_option=value_input_option) if value_input_option else self.worksheet.append_row(data)
                else:
                    self.worksheet.append_rows(data, value_input_option=value_input_option) if value_input_option else self.worksheet.append_rows(data)
                break
            except gspread.exceptions.APIError as e:
                print("Data write is temporarily paused due to exceeding API request limits. The system will resume operation after a short delay...")
                time.sleep(60)

    def open_sheet(self) -> gspread.Spreadsheet:
        if not self.gc:
            self._authorize()
        return self.gc.open_by_key(self.spreadsheet_id)

    def prepare_worksheet(self, sheet_name: str) -> tuple[gspread.Worksheet, int]:
        try:
            self.worksheet = self.open_sheet().worksheet(sheet_name)
            sheet_metadata = self.service.spreadsheets().get(spreadsheetId=self.spreadsheet_id).execute()
            sheets = sheet_metadata.get('sheets', [])
            self.sheet_id = next(
                sheet['properties']['sheetId']
                for sheet in sheets
                if sheet['properties']['title'] == sheet_name
            )
            print(f"Sheet {sheet_name} already exists.")
            print(f"Open the sheet at https://docs.google.com/spreadsheets/d/{self.spreadsheet_id}/edit#gid={self.sheet_id}")
            exit(0)
        except gspread.exceptions.WorksheetNotFound:
            try:
                self.open_sheet().add_worksheet(title=sheet_name, rows="1000", cols="26")
                self.worksheet = self.open_sheet().worksheet(sheet_name)
                sheet_metadata = self.service.spreadsheets().get(
                    spreadsheetId=self.spreadsheet_id).execute()
                sheets = sheet_metadata.get('sheets', [])
                self.sheet_id = next(
                    sheet['properties']['sheetId']
                    for sheet in sheets
                    if sheet['properties']['title'] == sheet_name
                )
                self.set_column_widths(start_col=0, end_col=26, width=185)
                return self.sheet_id
            except gspread.exceptions.APIError as e:
                print(f"Error creating sheet {sheet_name}: {e}")
                exit(1)
    
    def append_table_to_sheet(self, data: list[list[datetime | str]], current_date: str, found_users: int, start_row: int, num_rows: int = 96) -> None:
        self._safety_append_rows(data, value_input_option='USER_ENTERED')

        total_formula_row = []
        for col_index in range(2, found_users + 2):
            column_letter = get_column_letter(col_index)
            count_formula = f"COUNTIF({column_letter}{start_row}:{column_letter}{start_row + num_rows - 1}, \"<>\")"
            total_minutes = f"{count_formula} * 15"
            formatted_time_formula = f"=TEXT(INT({total_minutes} / 60), \"0\") & \":\" & TEXT(MOD({total_minutes}, 60), \"00\")"
            total_formula_row.append(formatted_time_formula)

        total_row = [[f'TOTAL [{current_date}]'] + total_formula_row]

        self._safety_append_rows(total_row, value_input_option='USER_ENTERED')

        start_border_row = start_row - 2  # 2 rows before the header row
        end_border_row = start_row + num_rows # 96 rows after the header row
        first_column_letter = 0
        last_column_letter = found_users + 1 # first column is the date
        self.table_formating(start_row=start_border_row, end_row=end_border_row, start_col=first_column_letter, end_col=last_column_letter)

    def append_all_totals(self, num_days: int, users_in_work: dict, start_date: str, stop_date: str) -> None:
        total_row_start = 3
        total_row_end = num_days * 99 + 1

        all_total_formula_row = []

        for col_index in range(2, len(users_in_work) + 2):
            column_letter = get_column_letter(col_index)

            all_total_minutes = f"SUM(ArrayFormula(VALUE(SPLIT(FILTER({column_letter}{total_row_start}:{column_letter}{total_row_end}, REGEXMATCH(A{total_row_start}:A{total_row_end}, \"TOTAL*\")), \":\")) * {{60, 1}}))"
            formatted_time_formula = f"=TEXT(INT({all_total_minutes} / 60), \"0\") & \":\" & TEXT(MOD({all_total_minutes}, 60), \"00\")"

            all_total_formula_row.append(formatted_time_formula)

        header_row = ['ALL TOTAL'] + list(users_in_work.keys())
        all_total_row = [f'{start_date} / {stop_date}'] + all_total_formula_row
        self._safety_append_rows([header_row, all_total_row], value_input_option='USER_ENTERED')

        start_border_row = total_row_end + 1 # 1 row after the last total row
        end_border_row = start_border_row + 2 # 2 rows after the last total row
        first_column_letter = 0
        last_column_letter = len(users_in_work) + 1 # first column is the date
        self.total_formating(start_row=start_border_row, end_row=end_border_row, start_col=first_column_letter, end_col=last_column_letter)

    def set_column_widths(self, start_col: int, end_col: int, width: int) -> None:
        requests = [
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": self.sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": start_col,
                        "endIndex": end_col
                    },
                    "properties": {
                        "pixelSize": width
                    },
                    "fields": "pixelSize"
                }
            }
        ]
        body = {'requests': requests}

        try:
            self.service.spreadsheets().batchUpdate(
                spreadsheetId=self.spreadsheet_id,
                body=body
            ).execute()
        except HttpError as err:
            print(f'An error occurred: {err}')

    def header_formating(self, start_row: int, start_col: int, end_col: int) -> None:
        red_color = self.hex_to_rgb("#AC3A4D")
        light_red_color = self.hex_to_rgb("#FFE3E6")

        requests = [
        # row formatting
        {
            "repeatCell": {
                "range": {
                    "sheetId": self.sheet_id,
                    "startRowIndex": start_row,
                    "endRowIndex": start_row + 1,
                    "startColumnIndex": start_col,
                    "endColumnIndex": end_col
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "LEFT",
                        "verticalAlignment": "MIDDLE",
                        "textFormat": {
                            "bold": True,
                            "fontSize": 20,
                            "foregroundColor": red_color
                        },
                        "backgroundColor": light_red_color,  
                    }
                },
                "fields": "userEnteredFormat(horizontalAlignment, verticalAlignment, textFormat, backgroundColor)"
            }
        }
        ]
        body = {'requests': requests}

        try:
            self.service.spreadsheets().batchUpdate(
                spreadsheetId=self.spreadsheet_id,
                body=body
            ).execute()
        except HttpError as err:
            print(f'An error occurred: {err}')
        
    def table_formating(self, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
        green_color = self.hex_to_rgb("#006100")
        light_green_color = self.hex_to_rgb("#DFF0E2")

        requests = [
        # header formatting
        {
            "repeatCell": {
                "range": {
                    "sheetId": self.sheet_id,
                    "startRowIndex": start_row,
                    "endRowIndex": start_row + 1,
                    "startColumnIndex": start_col,
                    "endColumnIndex": end_col
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                        "textFormat": {
                            "bold": True,
                            "fontSize": 13,
                            "foregroundColor": {
                                "red": 1.0,
                                "green": 1.0,
                                "blue": 1.0
                            }
                        },
                        "backgroundColor": green_color,  
                    }
                },
                "fields": "userEnteredFormat(horizontalAlignment, verticalAlignment, textFormat, backgroundColor)"
            }
        },
        # total formatting
        {
            "repeatCell": {
                "range": {
                    "sheetId": self.sheet_id,
                    "startRowIndex": end_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": start_col,
                    "endColumnIndex": end_col
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                        "textFormat": {
                            "bold": True,
                            "fontSize": 11,
                            "foregroundColor": green_color
                        },
                        "backgroundColor": light_green_color,  
                    }
                },
                "fields": "userEnteredFormat(horizontalAlignment, verticalAlignment, textFormat, backgroundColor)"
            }
        },
        # time slots formatting
        {
            "repeatCell": {
                "range": {
                    "sheetId": self.sheet_id,
                    "startRowIndex": start_row + 1,
                    "endRowIndex": end_row - 1,
                    "startColumnIndex": start_col,
                    "endColumnIndex": start_col + 1
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                        "textFormat": {
                            "bold": True,
                            "fontSize": 11,
                        },
                    }
                },
                "fields": "userEnteredFormat(horizontalAlignment, verticalAlignment, textFormat, backgroundColor)"
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": self.sheet_id,
                    "startRowIndex": start_row + 1,
                    "endRowIndex": end_row - 1,
                    "startColumnIndex": start_col + 1,
                    "endColumnIndex": end_col
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "LEFT",
                        "verticalAlignment": "MIDDLE",
                        "wrapStrategy": "CLIP",
                        "textFormat": {
                            "bold": False,
                            "fontSize": 11,
                        },
                    },
                },
                "fields": "userEnteredFormat(horizontalAlignment, verticalAlignment, textFormat, backgroundColor)"
            }
        },
        # all table borders
        {
            "repeatCell": {
                "range": {
                    "sheetId": self.sheet_id,
                    "startRowIndex": start_row,
                    "endRowIndex": end_row,
                    "startColumnIndex": start_col,
                    "endColumnIndex": end_col
                },
                "cell": {
                    "userEnteredFormat": {
                        "borders": {
                            "top": {"style": "SOLID"},
                            "bottom": {"style": "SOLID"},
                            "left": {"style": "SOLID"},
                            "right": {"style": "SOLID"}
                        }
                    }
                },
                "fields": "userEnteredFormat.borders"
            }
        }
        ]
        body = {'requests': requests}

        try:
            self.service.spreadsheets().batchUpdate(
                spreadsheetId=self.spreadsheet_id,
                body=body
            ).execute()
        except HttpError as err:
            print(f'An error occurred: {err}')

    def total_formating(self, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
        red_color = self.hex_to_rgb("#AC3A4D")
        light_red_color = self.hex_to_rgb("#FFE3E6")

        requests = [
        # header formatting
        {
            "repeatCell": {
                "range": {
                    "sheetId": self.sheet_id,
                    "startRowIndex": start_row,
                    "endRowIndex": start_row + 1,
                    "startColumnIndex": start_col,
                    "endColumnIndex": end_col
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                        "textFormat": {
                            "bold": True,
                            "fontSize": 13,
                            "foregroundColor": {
                                "red": 1.0,
                                "green": 1.0,
                                "blue": 1.0
                            }
                        },
                        "backgroundColor": red_color,  
                    }
                },
                "fields": "userEnteredFormat(horizontalAlignment, verticalAlignment, textFormat, backgroundColor)"
            }
        },
        # total formatting
        {
            "repeatCell": {
                "range": {
                    "sheetId": self.sheet_id,
                    "startRowIndex": end_row - 1,
                    "endRowIndex": end_row,
                    "startColumnIndex": start_col,
                    "endColumnIndex": end_col
                },
                "cell": {
                    "userEnteredFormat": {
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                        "textFormat": {
                            "bold": True,
                            "fontSize": 11,
                            "foregroundColor": red_color
                        },
                        "backgroundColor": light_red_color,  
                    }
                },
                "fields": "userEnteredFormat(horizontalAlignment, verticalAlignment, textFormat, backgroundColor)"
            }
        },
        # all table borders
        {
            "repeatCell": {
                "range": {
                    "sheetId": self.sheet_id,
                    "startRowIndex": start_row,
                    "endRowIndex": end_row,
                    "startColumnIndex": start_col,
                    "endColumnIndex": end_col
                },
                "cell": {
                    "userEnteredFormat": {
                        "borders": {
                            "top": {"style": "SOLID"},
                            "bottom": {"style": "SOLID"},
                            "left": {"style": "SOLID"},
                            "right": {"style": "SOLID"}
                        }
                    }
                },
                "fields": "userEnteredFormat.borders"
            }
        }
        ]
        body = {'requests': requests}

        try:
            self.service.spreadsheets().batchUpdate(
                spreadsheetId=self.spreadsheet_id,
                body=body
            ).execute()
        except HttpError as err:
            print(f'An error occurred: {err}')

    def hex_to_rgb(self, hex_color: str) -> dict:
        hex_color = hex_color.lstrip('#')
        if len(hex_color) == 6:
            r, g, b = hex_color[0:2], hex_color[2:4], hex_color[4:6]
        elif len(hex_color) == 8:
            r, g, b = hex_color[2:4], hex_color[4:6], hex_color[6:8]
        else:
            raise ValueError("Invalid hex color format")
        return {
            "red": int(r, 16) / 255.0,
            "green": int(g, 16) / 255.0,
            "blue": int(b, 16) / 255.0
        }
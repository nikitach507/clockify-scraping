from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

column_widths = {1: 20.0, 2: 20.0}

def set_column_widths(worksheet: Worksheet, max_col: int, widths: dict[int, float] = column_widths) -> None:
    for col in range(1, max_col + 1):
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].width = widths[2]

def apply_styles(worksheet: Worksheet, min_row: int, max_row: int, min_col: int, max_col: int, 
                  fill_color: str = None, horizontal_alignment: str = None, vertical_alignment: str = None, 
                  border: Border = thin_border) -> None:
    """
    Applies styling to a range of cells in the worksheet.
    """

    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid") if fill_color else None
    alignment = Alignment(horizontal=horizontal_alignment, vertical=vertical_alignment, wrapText=False) if horizontal_alignment or vertical_alignment else Alignment(wrapText=False)
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if fill:
                cell.fill = fill
            if alignment:
                if not cell.value or len(str(cell.value)) <= 15:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=False)
                else:
                    cell.alignment = alignment
            if border:
                cell.border = border
            else:
                continue
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
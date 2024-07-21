import re
import click
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

column_widths = {
        1: 12.0,
        2: 20.0,
    }

def set_column_widths(worksheet: Worksheet, widths: dict[int, float], max_col: int) -> None:
    worksheet.column_dimensions['A'].width = widths[1]

    for col in range(2, max_col + 1):
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].width = widths[2]
    

def generate_total_rows(current_date: str, start_row: int, number_users: int, num_rows: int = 96) -> list[list[str]]:
    total_formula_row = []
    for col_index in range(2, number_users+2):
        column_letter = get_column_letter(col_index)

        count_formula = f"COUNTIF({column_letter}{start_row}:{column_letter}{start_row + num_rows - 1}, \"<>\")"
        total_minutes = f"{count_formula} * 15"
        formatted_time_formula = f"=TEXT(INT({total_minutes} / 60), \"0\") & \":\" & TEXT(MOD({total_minutes}, 60), \"00\")"
        total_formula_row.append(formatted_time_formula)

    total_row = [f'TOTAL [{current_date}]'] + total_formula_row
    return [total_row]

def append_data_to_sheet(worksheet: Worksheet, data: list[list[str]], current_date: str, number_users: int, start_row: int) -> None:
    for row in data:
        worksheet.append(row)

    total_rows = generate_total_rows(current_date, start_row, number_users)
    for total_row in total_rows:
        worksheet.append(total_row)

    green_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    wrap_alignment = Alignment(wrapText=True, vertical='center')


    for cell in worksheet[worksheet.max_row]:
        cell.fill = green_fill

    for row in worksheet.iter_rows(min_row=start_row-1, max_row=worksheet.max_row, min_col=1, max_col=number_users + 1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = wrap_alignment
            if cell.row == worksheet.max_row:
                if cell.column != 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    set_column_widths(worksheet, column_widths, number_users+1)

def generate_all_totals(number_users: int, num_days: int, start_date: str, stop_date: str) -> tuple[list[list[str]], list[dict]]:
    total_row_end = num_days * 99 - 1
    iteration = 98
    buffer_30_days = 0
    buffer_rows = []
    all_total_formula_row = []
    all_buffers_rows = []

    for col_index in range(2, number_users+2):
        column_letter = get_column_letter(col_index)
        total_minutes = ""
        while iteration <= total_row_end:
            formula = f"HOUR(TIMEVALUE({column_letter}{iteration}))*60+MINUTE(TIMEVALUE({column_letter}{iteration}))"
            if buffer_30_days < 60:
                total_minutes += formula
                if iteration != total_row_end:
                    total_minutes = total_minutes + "+"
                buffer_30_days += 1
                iteration += 99
            else:
                buffer_rows.append(total_minutes[:-1])
                total_minutes = ""
                buffer_30_days = 0
        
        buffer_rows.append(total_minutes)

        print(f"{buffer_rows} \n")
        
        join_minutes = ""

        for number in range(0, len(buffer_rows)):
            join_minutes += f"{column_letter}{total_row_end + number + 2 + 2}" + "+" if number != len(buffer_rows) - 1 else f"{column_letter}{total_row_end + number + 2 + 2}"

        print(f"={join_minutes} \n")
        total_formula = f"=TEXT(INT(({join_minutes}) / 60), \"0\") & \":\" & TEXT(MOD(({join_minutes}), 60), \"00\")"
        all_total_formula_row.append(total_formula)
        print(f"{total_formula} \n")

        buffer_dict = {}
        for number, row in enumerate(buffer_rows):
            
            buffer_dict[number] = row

        all_buffers_rows.append([buffer_dict])

        iteration = 98
        buffer_rows = []
        buffer_30_days = 0
        join_minutes = ""

    print(f"{all_total_formula_row} \n ")

    all_total_row = [f'ALL TOTAL [{start_date} / {stop_date}]'] + all_total_formula_row

    return [all_total_row], all_buffers_rows
    

def append_all_totals(worksheet: Worksheet, num_days: int, number_users: int, start_date: str, stop_date: str) -> None:
    all_totals, all_buffer = generate_all_totals(number_users, num_days, start_date, stop_date)
    print(f"All totals: {all_totals} \n")
    print(f"All buffer: {all_buffer} \n")
    for row in all_totals:
        worksheet.append(row) 

    green_fill = PatternFill(start_color="DDD9C4", end_color="DDD9C4", fill_type="solid")

    for cell in worksheet[worksheet.max_row]:
        cell.fill = green_fill

    for row in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row, min_col=1, max_col=number_users + 1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrapText=True)
            if cell.column != 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
    worksheet.append(["·"])

    buffer_row = []
    for number in range(0, len(all_buffer[0][0].keys())):
        buffer_row.append(f"Buffer 60 days")
        for buffer in all_buffer:
            print(f"Buffer: {buffer[0][number]}")
            buffer_row.append("=" + buffer[0][number])
        
        worksheet.append(buffer_row)
        buffer_row = []